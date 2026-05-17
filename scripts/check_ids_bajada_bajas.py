#!/usr/bin/env python3
"""
Corrobora access IDs contra las mismas fuentes que la consulta índice:

1. ``aux.bajada_inventario`` (detalle tipo bajada / histórico de bajada).
2. ``aux.bajas_de_inventario`` y ``aux.bajas_inventario`` (bajas explícitas).
3. Inventario FAT activo (``access_id_topologia``).

Por defecto exporta CSV **separado por punto y coma** (``;``), UTF-8 con BOM y línea
``sep=;`` para que Excel (región ES/LATAM) abra columnas y acentos bien. Para una
hoja con formato, usá ``--out resultado.xlsx``.

Usa **una conexión** y ``clasificar_access_id_bajada_bajas_fat_cur`` por ID.

Requiere variables de entorno / ``.env`` de la app (DATABASE_URL o DB_*) como el resto del proyecto.

Ejemplos::

    python scripts/check_ids_bajada_bajas.py lista.txt --out resultado.xlsx
    python scripts/check_ids_bajada_bajas.py lista.txt --out resultado.csv
    python scripts/check_ids_bajada_bajas.py lista.txt --raw > datos_tecnicos.csv
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from db import db_cursor  # noqa: E402
from services.inventory import clasificar_access_id_bajada_bajas_fat_cur  # noqa: E402

_CATEGORIA_POR_TABLA: dict[str, str] = {
    "aux.bajada_inventario": "Bajada (histórico / detalle)",
    "aux.bajas_de_inventario": "Baja explícita (tabla bajas_de_inventario)",
    "aux.bajas_inventario": "Baja explícita (tabla bajas_inventario)",
    "ninguna": "Sin fila en tablas aux",
}

# Texto solo ASCII en CSV: si Excel abre como ANSI, no se rompe la lectura.
_CATEGORIA_CSV_ASCII: dict[str, str] = {
    "aux.bajada_inventario": "Bajada (historico / detalle)",
    "aux.bajas_de_inventario": "Baja explicita (tabla bajas_de_inventario)",
    "aux.bajas_inventario": "Baja explicita (tabla bajas_inventario)",
    "ninguna": "Sin fila en tablas aux",
}

_FIELDNAMES_RAW = [
    "access_id",
    "tabla_aux",
    "bajada_inventario",
    "bajas_tabla",
    "inventario_fat_activo",
    "ubicacion_resumen",
]

_FIELDNAMES_PRETTY = [
    "Access ID",
    "Categoria",
    "Tabla en base de datos",
    "Activo en inventario FAT",
]


def _norm_id(line: str) -> str | None:
    s = line.strip()
    if not s or s.lower() in ("access-id", "access_id", "aid", "id"):
        return None
    if re.fullmatch(r"\d+", s):
        return s
    return None


def load_ids(path: Path) -> list[str]:
    raw = path.read_text(encoding="utf-8", errors="replace").splitlines()
    out: list[str] = []
    seen: set[str] = set()
    for line in raw:
        aid = _norm_id(line)
        if not aid or aid in seen:
            continue
        seen.add(aid)
        out.append(aid)
    return out


def _fila_pretty(row: dict[str, str], *, for_xlsx: bool = False) -> dict[str, str]:
    tabla = row.get("tabla_aux") or "ninguna"
    cat = (_CATEGORIA_POR_TABLA if for_xlsx else _CATEGORIA_CSV_ASCII).get(tabla, tabla)
    tabla_sql = tabla if tabla != "ninguna" else "(ninguna de las tablas aux consultadas)"
    if for_xlsx:
        fat = "Sí" if row.get("inventario_fat_activo") == "si" else "No"
    else:
        fat = "Si" if row.get("inventario_fat_activo") == "si" else "No"
    return {
        "Access ID": row.get("access_id") or "",
        "Categoria": cat,
        "Tabla en base de datos": tabla_sql,
        "Activo en inventario FAT": fat,
    }


_XLSX_HEADERS = [
    "Access ID",
    "Categoría",
    "Tabla en base de datos",
    "¿Activo en inventario FAT?",
]


def _write_csv_rows(
    path: Path,
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    encoding: str,
    delimiter: str = ",",
    excel_sep_hint: str | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding=encoding, newline="") as f:
        if excel_sep_hint:
            f.write(excel_sep_hint)
        w = csv.DictWriter(
            f,
            fieldnames=fieldnames,
            delimiter=delimiter,
            lineterminator="\n",
            extrasaction="ignore",
        )
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(path: Path, rows_raw: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Falta openpyxl. Instalá dependencias: pip install openpyxl"
        ) from e

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Clasificación AID"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, title in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r, row in enumerate(rows_raw, start=2):
        pretty = _fila_pretty(row, for_xlsx=True)
        ws.cell(row=r, column=1, value=str(pretty["Access ID"])).number_format = "@"
        ws.cell(row=r, column=2, value=pretty["Categoria"])
        ws.cell(row=r, column=3, value=pretty["Tabla en base de datos"])
        ws.cell(row=r, column=4, value=pretty["Activo en inventario FAT"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_XLSX_HEADERS))}{len(rows_raw) + 1}"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 26

    wb.save(path)


def _write_pretty_csv_to_stdout(
    pretty_excel_sep: str,
    stdout_fn: list[str],
    stdout_rows: list[dict[str, str]],
    pretty_csv_delim: str,
) -> None:
    """CSV legible por stdout: intenta UTF-8 + BOM; si el terminal es cp1252, omite el BOM."""
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (OSError, ValueError, AttributeError, TypeError):
        pass
    try:
        sys.stdout.write("\ufeff" + pretty_excel_sep)
    except UnicodeEncodeError:
        sys.stdout.write(pretty_excel_sep)
    w = csv.DictWriter(
        sys.stdout,
        fieldnames=stdout_fn,
        delimiter=pretty_csv_delim,
        lineterminator="\n",
        extrasaction="ignore",
    )
    w.writeheader()
    for r in stdout_rows:
        w.writerow(r)
    sys.stdout.flush()


def main() -> int:
    p = argparse.ArgumentParser(description="Verificar access IDs en bajada / bajas aux.")
    p.add_argument(
        "archivo",
        type=Path,
        help="Archivo de texto con un access_id numérico por línea (p. ej. export DTV).",
    )
    p.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="Copia adicional del export (mismo formato que --out o CSV legible por defecto).",
    )
    p.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Archivo de salida: .xlsx (recomendado) o .csv (UTF-8 BOM, separador ;, linea sep=; para Excel ES).",
    )
    p.add_argument(
        "--raw",
        action="store_true",
        help="CSV técnico: nombres de columnas internos (access_id, tabla_aux, …), sin BOM.",
    )
    p.add_argument(
        "--progress",
        type=int,
        default=50,
        metavar="N",
        help="Cada N IDs imprime avance en stderr (0 = desactivar). Por defecto: 50.",
    )
    args = p.parse_args()
    if not args.archivo.is_file():
        print(f"No existe el archivo: {args.archivo}", file=sys.stderr, flush=True)
        return 2

    ids = load_ids(args.archivo)
    if not ids:
        print("No se encontraron IDs numéricos en el archivo.", file=sys.stderr, flush=True)
        return 1

    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(line_buffering=True)
        except (OSError, ValueError, AttributeError):
            pass

    out_path = args.out
    use_xlsx = out_path is not None and out_path.suffix.lower() == ".xlsx"
    pretty = not args.raw

    out_f = None
    out_cm: csv.DictWriter | None = None
    pretty_csv_delim = ";"
    pretty_excel_sep = "sep=;\n"
    if out_path is not None and not use_xlsx:
        enc = "utf-8" if args.raw else "utf-8-sig"
        fn = _FIELDNAMES_RAW if args.raw else _FIELDNAMES_PRETTY
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_f = out_path.open("w", encoding=enc, newline="")
        if pretty:
            out_f.write(pretty_excel_sep)
        delim = "," if args.raw else pretty_csv_delim
        out_cm = csv.DictWriter(
            out_f,
            fieldnames=fn,
            delimiter=delim,
            lineterminator="\n",
            extrasaction="ignore",
        )
        out_cm.writeheader()
        out_f.flush()

    if pretty:
        stdout_fn = _FIELDNAMES_PRETTY
        stdout_enc_hint = "UTF-8 con BOM; separador ; y linea sep=; (ideal para > archivo.csv)"
    else:
        stdout_fn = _FIELDNAMES_RAW
        stdout_enc_hint = "utf-8, separador coma"

    stdout_rows: list[dict[str, str]] = []

    if out_path is None:
        print(
            f"Salida por stdout: columnas {'legibles' if pretty else 'técnicas'} ({stdout_enc_hint}).",
            file=sys.stderr,
            flush=True,
        )

    counts: dict[str, int] = {}
    rows_kept: list[dict[str, str]] = []
    t0 = time.perf_counter()
    n = len(ids)
    prog_txt = f"cada {args.progress}" if args.progress else "sin avance intermedio"
    print(f"Procesando {n} access_id (una conexión DB, {prog_txt})…", file=sys.stderr, flush=True)

    with db_cursor() as cur:
        for i, aid in enumerate(ids, start=1):
            row = clasificar_access_id_bajada_bajas_fat_cur(cur, aid)
            rows_kept.append(row)
            u = row["tabla_aux"]
            counts[u] = counts.get(u, 0) + 1

            if out_cm is not None and out_f is not None:
                out_row = row if not pretty else _fila_pretty(row, for_xlsx=False)
                out_cm.writerow(out_row)
                out_f.flush()
            elif out_path is None:
                stdout_rows.append(row if not pretty else _fila_pretty(row, for_xlsx=False))

            if args.progress and i % args.progress == 0:
                elapsed = time.perf_counter() - t0
                print(f"  … {i}/{n}  ({elapsed:.0f}s)", file=sys.stderr, flush=True)

    if out_f is not None:
        out_f.close()

    if use_xlsx:
        _write_xlsx(out_path, rows_kept)

    if out_path is None:
        if pretty:
            _write_pretty_csv_to_stdout(pretty_excel_sep, stdout_fn, stdout_rows, pretty_csv_delim)
        else:
            w = csv.DictWriter(
                sys.stdout,
                fieldnames=stdout_fn,
                delimiter=",",
                lineterminator="\n",
                extrasaction="ignore",
            )
            w.writeheader()
            for r in stdout_rows:
                w.writerow(r)
            sys.stdout.flush()

    elapsed = time.perf_counter() - t0
    print(f"Listo en {elapsed:.1f}s. Resumen por tabla_aux:", file=sys.stderr, flush=True)
    for k in sorted(counts.keys(), key=lambda x: (-counts[x], x)):
        print(f"  {k}: {counts[k]}", file=sys.stderr, flush=True)

    if args.csv is not None:
        if args.raw:
            _write_csv_rows(args.csv, rows_kept, _FIELDNAMES_RAW, encoding="utf-8", delimiter=",")
        else:
            pretty_rows = [_fila_pretty(r, for_xlsx=False) for r in rows_kept]
            _write_csv_rows(
                args.csv,
                pretty_rows,
                _FIELDNAMES_PRETTY,
                encoding="utf-8-sig",
                delimiter=pretty_csv_delim,
                excel_sep_hint=pretty_excel_sep,
            )
        print(f"Copia: {args.csv}", file=sys.stderr, flush=True)

    if out_path is not None:
        print(
            f"Archivo principal: {out_path} ({'Excel' if use_xlsx else 'CSV'})",
            file=sys.stderr,
            flush=True,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
